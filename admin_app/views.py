from django.shortcuts import render, get_object_or_404, redirect
from Scholarshipapp.models import *
from django.db import IntegrityError
from datetime import datetime, timedelta
from django.core.files.storage import FileSystemStorage
from django.utils.datastructures import MultiValueDictKeyError
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login
from django.http import HttpResponse
from django.contrib import messages
from Scholarshipapp.forms import *


# Create your views here..
import openpyxl
from io import BytesIO
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.utils import get_column_letter



def export_scholarship_details_to_excel(request):
    scholarship_details = Scholarship_details.objects.all()

    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Scholarship Details"

    # Define the headers
    headers = [
        'Scholarship Name', 'Offered By', 'Award', 'Category', 'Education', 
        'All Scholarship', 'States', 'Published Date', 'Deadline', 'Description'
    ]

    # Write the headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f'{column_letter}1'] = header

    # Write the data rows
    for row_num, detail in enumerate(scholarship_details, 2):
        worksheet[f'A{row_num}'] = detail.Scholarship_name
        worksheet[f'B{row_num}'] = detail.Offered_by
        worksheet[f'C{row_num}'] = detail.Award
        worksheet[f'D{row_num}'] = ', '.join([cat.S_Category for cat in detail.S_Category.all()])
        worksheet[f'E{row_num}'] = ', '.join([edu.Education for edu in detail.Education.all()])
        worksheet[f'F{row_num}'] = ', '.join([sch.All_Scholarship for sch in detail.All_Scholarship.all()])
        worksheet[f'G{row_num}'] = ', '.join([state.States for state in detail.States.all()])
        worksheet[f'H{row_num}'] = detail.Published_date.strftime('%Y-%m-%d') if detail.Published_date else ''
        worksheet[f'I{row_num}'] = detail.Dead_line.strftime('%Y-%m-%d') if detail.Dead_line else ''
        worksheet[f'J{row_num}'] = detail.Scholar_Description

    # Create an in-memory buffer to store the workbook
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # Create an HTTP response with the appropriate content type for Excel
    response = HttpResponse(
        content=buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=scholarship_details.xlsx'

    return response



# def admin_register(request):
#     if request.method == 'POST':
#         username = request.POST['username']
#         password = request.POST['password']
#         if User.objects.filter(username=username).exists():
#             messages.error(request, 'Username already exists')
#         else:
#             user = User.objects.create_user(username=username, password=password)
#             user.save()
#             messages.success(request, 'Registration successful')
#             return redirect('admin_app:admin_login')
#     return render(request, 'admin_register.html')

def admin_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('admin_app:admindex')
        else:
            messages.error(request, 'Invalid username or password')
    return render(request, 'admin_login.html')
    
    
def admindex(request):
    total_registered_users = Register.objects.count()
    total_applied_scholarships = applied_scholarship.objects.count()
    
    context = {
        'total_registered_users': total_registered_users,
        'total_applied_scholarships': total_applied_scholarships,
    }
    
    return render(request, 'admindex.html', context)





def scholar_det_form(request):
    s_category = S_Category.objects.all().order_by('S_Category')
    states = States.objects.all().order_by('States')
    education = Education.objects.all().order_by('Education')
    all_Scholarship = All_Scholarship.objects.all().order_by('All_Scholarship')
    context = {'s_category': s_category, 'states': states,
               'all_Scholarship': all_Scholarship, 'education': education}

    if request.method == 'POST':
        scholarship_logo = request.FILES.get('scholarship_logo')
        scholarship_name = request.POST.get('scholarship_name')
        offered_by = request.POST.get('offered_by')
        # plan_status = request.POST.get('plan_status')
        award = request.POST.get('award')
        published_date = request.POST.get('published_date')
        dead_line = request.POST.get('dead_line')
        scholar_description = request.POST.get('scholar_description')

        # Get selected values for many-to-many fields
        s_category_ids = request.POST.getlist('s_category[]')
        education_ids = request.POST.getlist('education[]')
        all_scholarship_ids = request.POST.getlist('all_scholarship[]')
        states_ids = request.POST.getlist('states[]')

        # Create Scholarship_details instance
        scholarship = Scholarship_details.objects.create(
            Scholarship_logo=scholarship_logo,
            Scholarship_name=scholarship_name,
            Offered_by=offered_by,
            # Plan_status=plan_status,
            Award=award,
            Published_date=published_date,
            Dead_line=dead_line,
            Scholar_Description=scholar_description
        )

        # Add selected values to many-to-many fields
        scholarship.S_Category.add(*s_category_ids)
        scholarship.Education.add(*education_ids)
        scholarship.All_Scholarship.add(*all_scholarship_ids)
        scholarship.States.add(*states_ids)

        # Redirect to a success page
        return redirect('admin_app:scholar_det_form')

    else:
        # Render the form page
        return render(request, 'scholar_det_form.html', context)


def scholarship_detail_table(request):
    scholarship_details = Scholarship_details.objects.all()
    context = {'scholarship_details': scholarship_details}
    return render(request, 'scholarship_details_tables.html', context)

def edit_scholarship_detail(request, pk):
    scholarship_detail = get_object_or_404(Scholarship_details, pk=pk)
    if request.method == 'POST':
        form = ScholarshipForm(request.POST, request.FILES, instance=scholarship_detail)
        if form.is_valid():
            form.save()
            return redirect('admin_app:scholarship_detail_table')
    else:
        form = ScholarshipForm(instance=scholarship_detail)
    return render(request, 'edit_scholarship_detail.html', {'form': form})
    
def delete_scholarship_detail(request, pk):
    scholarship = get_object_or_404(Scholarship_details, pk=pk)
    if request.method == 'POST':
        scholarship.delete()
        return redirect('admin_app:scholarship_detail_table')
    return render(request, 'delete_scholarship_detail.html', {'scholarship': scholarship})

# def scholar_det_form(request):
#     s_category=S_Category.objects.all()
#     all_Scholarship=All_Scholarship.objects.all()
#     states=States.objects.all()
#     education=Education.objects.all()
#     context={'s_category':s_category, 'states':states,'education':education,'all_Scholarship':all_Scholarship}
#     if request.method=='POST':
#         Scholarship_logo=request.POST['scholarship_logo']
#         Scholarship_name=request.POST['scholarship_name']
#         Offered_by=request.POST['offered_by']
#         Plan_status=request.POST['plan_status']
#         Award=request.POST['award']

#         s_category=request.POST['s_category']
#         S_Category_selected=S_Category.objects.get(S_Category=s_category)

#         education=request.POST['education']
#         Education_selected=Education.objects.get(Education=education)

#         states=request.POST['states']
#         States_selected=States.objects.get(States=states)

#         all_Scholarship=request.POST['all_scholarship']
#         All_Scholarship_selected=request.POST['all_scholarship']

#         Time_field=request.POST['time_field']
#         Published_date=request.POST['published_date']
#         Dead_line=request.POST['dead_line']
#         Scholar_Description=request.POST['scholar_description']
#         print(S_Category_selected)
#         # Register.objects.create(u_name=u_name,f_Name=f_Name,l_Name=l_Name,dob=dob,Mob=Mob,Email=Email,Category=Category_selected,Education=Education_selected,Password=Password,Re_Password=Re_Password,States=States_selected)
#     return render(request,'scholar_det_form.html',context)


def eligibility_form(request):
    if request.method == 'POST':
        plan_name = request.POST['plan_name']
        plan_price = request.POST['plan_price']
        plan_image = None 
        # Check if a file was uploaded
        if 'plan_image' in request.FILES:
            plan_image = request.FILES['plan_image']
        
        User_Plan.objects.create(plan_name=plan_name, plan_price=plan_price, plan_image=plan_image)
        return redirect('admin_app:eligibility_form_view')
    
    return render(request, 'eligibility_form.html')
    
def eligibility_form_view(request):
    all_plans = User_Plan.objects.all()
    context = {'all_plans': all_plans}
    return render(request, 'eligibility_list.html', context)

def edit_eligibility_form(request, pk):
    plan = get_object_or_404(User_Plan, pk=pk)
    if request.method == 'POST':
        plan.plan_name = request.POST['plan_name']
        plan.plan_price = request.POST['plan_price']
        if request.FILES.get('plan_image'):
            plan.plan_image = request.FILES['plan_image']
        plan.save()
        return redirect('admin_app:eligibility_form_view')
    context = {'plan': plan}
    return render(request, 'edit_eligibility_form.html', context)

def delete_eligibility_form(request, pk):
    plan = get_object_or_404(User_Plan, pk=pk)
    plan.delete()
    return redirect('admin_app:eligibility_form_view')

# def user_tables(request):
#     if request.method == 'POST':
#         scholarship_id = request.POST.get('scholarship_id')
#         action = request.POST.get('action')
#         scholarship = applied_scholarship.objects.get(id=scholarship_id)
#         approval, created = Approval.objects.get_or_create(user=scholarship.user, applied_scholarship=scholarship)
#         approval.status = action
#         approval.save()
#         return redirect('admin_app:user_tables')

#     users = Register.objects.all()
#     user_profiles = User_Profile.objects.all()
#     applied_scholarships = applied_scholarship.objects.all()

#     user_scholarship_data = []

#     for user in users:
#         profile = next((profile for profile in user_profiles if profile.user_id == user.id), None)
#         scholarships = [app_sch for app_sch in applied_scholarships if app_sch.user_id == user.id]

#         for scholarship in scholarships:
#             user_scholarship_data.append({
#                 'user': user,
#                 'profile': profile,
#                 'scholarship': scholarship,
#             })

#     context = {
#         'user_scholarship_data': user_scholarship_data,
#     }
#     return render(request, 'user_tables.html', context)


def user_tables(request):
    if request.method == 'POST':
        scholarship_id = request.POST.get('scholarship_id')
        action = request.POST.get('action')
        scholarship = applied_scholarship.objects.get(id=scholarship_id)
        approval, created = Approval.objects.get_or_create(
            user=scholarship.user, applied_scholarship=scholarship)
        approval.status = action
        approval.save()

    users = Register.objects.all()
    user_profiles = User_Profile.objects.all()
    applied_scholarships = applied_scholarship.objects.all()

    user_scholarship_data = []

    for user in users:
        profile = next(
            (profile for profile in user_profiles if profile.user_id == user.id), None)
        scholarships = [
            app_sch for app_sch in applied_scholarships if app_sch.user_id == user.id]

        for scholarship in scholarships:
            try:
                approval = Approval.objects.get(
                    user=user, applied_scholarship=scholarship)
            except Approval.DoesNotExist:
                approval = None
            user_scholarship_data.append({
                'user': user,
                'profile': profile,
                'scholarship': scholarship,
                'approval': approval,
            })

    context = {
        'user_scholarship_data': user_scholarship_data,
    }
    return render(request, 'user_tables.html', context)


def commeneted(request):
    s_category = S_Category.objects.all()
    context = {'s_category': s_category}
    return render(request, 'commeneted.html', context)
    
def scolar_disply(request):
    all_scholarships = All_Scholarship.objects.all()
    context = {'all_scholarships': all_scholarships}
    return render(request, 'scolar_disply.html', context)

def edit_scholarship(request, pk):
    scholarship = get_object_or_404(All_Scholarship, pk=pk)
    if request.method == 'POST':
        scholarship_name = request.POST.get('scholarship_name')
        scholarship.All_Scholarship = scholarship_name
        scholarship.save()
        return redirect('admin_app:scolar_disply')
    
    context = {
        'scholarship': scholarship
    }
    return render(request, 'edit_scholarship.html', context)


def confirm_delete(request, pk):
    instance = get_object_or_404(All_Scholarship, pk=pk)
    instance.delete()
    return redirect('admin_app:scolar_disply')

def add_scholarship(request):
    if request.method == 'POST':
        scholarship_name = request.POST.get('scholarship_name')
        new_scholarship = All_Scholarship(All_Scholarship=scholarship_name)
        new_scholarship.save()
        return redirect('admin_app:scolar_disply')
    return render(request, 'scolar_disply.html')

   
def category_list(request):
    all_category = S_Category.objects.all()
    context = {'all_category': all_category}
    return render(request, 'category_list.html', context)

def edit_category(request, pk):
    category = get_object_or_404(S_Category, pk=pk)
    if request.method == 'POST':
        category_name = request.POST.get('category_name')
        category.S_Category = category_name
        category.save()
        return redirect('admin_app:category_list')
    
    context = {'category': category}
    return render(request, 'edit_cat.html', context)

def category_delete(request, pk):
    instance = get_object_or_404(S_Category, pk=pk)
    instance.delete()
    return redirect('admin_app:category_list')

def add_category(request):
    if request.method == 'POST':
        category_name = request.POST.get('category_name')
        new_category = S_Category(S_Category=category_name)
        new_category.save()
        return redirect('admin_app:category_list')
    return render(request, 'category_list.html')

def states_list(request):
    all_states = States.objects.all()
    context = {'all_states': all_states}
    return render(request, 'states_list.html', context)

def add_state(request):
    if request.method == 'POST':
        state_name = request.POST.get('state_name')
        new_state = States(States=state_name)
        new_state.save()
        return redirect('admin_app:states_list')
    return render(request, 'states_list.html')

def edit_states(request, pk):
    states = get_object_or_404(States, pk=pk)
    if request.method == 'POST':
        states_name = request.POST.get('states_name')
        states.States = states_name
        states.save()
        return redirect('admin_app:states_list')
    
    context = {'states': states}
    return render(request, 'edit_stat.html', context)

def states_delete(request, pk):
    instance = get_object_or_404(States, pk=pk)
    instance.delete()
    return redirect('admin_app:states_list')



def education_list(request):
    education_queryset = Education.objects.all()
    context = {'education_queryset': education_queryset} 
    return render(request, 'education_list.html', context)

def edit_education(request, pk):
    education = get_object_or_404(Education, pk=pk)  
    if request.method == 'POST':
        education_name = request.POST.get('education_name')
        education.Education = education_name
        education.save()
        return redirect('admin_app:education_list')
    
    context = {'education': education}
    return render(request, 'edit_education.html', context)

def education_delete(request, pk):
    instance = get_object_or_404(Education, pk=pk)
    instance.delete()
    return redirect('admin_app:education_list')

def add_education(request):
    if request.method == 'POST':
        education_name = request.POST.get('education_name')
        new_education = Education(Education=education_name)
        new_education.save()
        return redirect('admin_app:education_list')
    return render(request, 'education_list.html')


def exam_calender_view(request):
    if request.method == 'POST':
        xam_name = request.POST['xam_name']
        xam_date = request.POST['xam_date']
        
        try:
            xam_date = datetime.strptime(xam_date, '%Y-%m-%d').date()
            exam_calender.objects.create(xam_name=xam_name, xam_date=xam_date)
            return redirect('admin_app:exam_calender_view')
        except ValueError:
            return render(request, 'exam_calender.html', {'error': 'Invalid date format. Please use YYYY-MM-DD.'})

    all_exam = exam_calender.objects.all()
     
    context = {'all_exam': all_exam}
    
    return render(request, 'exam_calender.html', context)


def edit_exam(request, pk):
    exam = get_object_or_404(exam_calender, pk=pk)  
    if request.method == 'POST':
        xam_name = request.POST.get('xam_name')
        xam_date = request.POST.get('xam_date')
        try:
            xam_date = datetime.strptime(xam_date, '%Y-%m-%d').date()
            exam.xam_name = xam_name
            exam.xam_date = xam_date
            exam.save()
            return redirect('admin_app:exam_calender_view')
        except ValueError:
            return render(request, 'edit_exam.html', {'exam': exam, 'error': 'Invalid date format. Please use YYYY-MM-DD.'})

    context = {'exam': exam}
    return render(request, 'edit_exam.html', context)

def exam_delete(request, pk):
    instance = get_object_or_404(exam_calender, pk=pk)
    instance.delete()
    return redirect('admin_app:exam_calender_view')